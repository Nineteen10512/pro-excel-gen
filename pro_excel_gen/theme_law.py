from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .themes import ThemeProfile, get_theme


MIN_CONTRAST_RATIO = 4.5


def _normalize_hex(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip().replace("#", "")
    if len(value) == 8:
        value = value[-6:]
    if len(value) != 6:
        return None
    return f"#{value.upper()}"


def _rgb(hex_color: str) -> tuple[float, float, float]:
    color = _normalize_hex(hex_color) or "#000000"
    return tuple(int(color[idx : idx + 2], 16) / 255 for idx in (1, 3, 5))


def _linear(channel: float) -> float:
    if channel <= 0.03928:
        return channel / 12.92
    return ((channel + 0.055) / 1.055) ** 2.4


def contrast_ratio(foreground: str, background: str) -> float:
    fg = _rgb(foreground)
    bg = _rgb(background)
    lum_fg = 0.2126 * _linear(fg[0]) + 0.7152 * _linear(fg[1]) + 0.0722 * _linear(fg[2])
    lum_bg = 0.2126 * _linear(bg[0]) + 0.7152 * _linear(bg[1]) + 0.0722 * _linear(bg[2])
    lighter = max(lum_fg, lum_bg)
    darker = min(lum_fg, lum_bg)
    return (lighter + 0.05) / (darker + 0.05)


def theme_palette(theme: str | ThemeProfile | None = None) -> set[str]:
    profile = theme if isinstance(theme, ThemeProfile) else get_theme(theme)
    return {
        profile.accent.upper(),
        profile.accent_2.upper(),
        profile.accent_3.upper(),
        profile.background.upper(),
        profile.header_bg.upper(),
        profile.header_fg.upper(),
        profile.grid.upper(),
        profile.good.upper(),
        profile.warn.upper(),
        profile.bad.upper(),
        "#000000",
        "#FFFFFF",
    }


def _cell_colors(cell, profile: ThemeProfile) -> tuple[str, str]:
    bg = profile.background
    if cell.fill and cell.fill.fill_type == "solid":
        bg = _normalize_hex(cell.fill.fgColor.rgb) or bg
    fg = "#000000"
    if cell.font and cell.font.color and cell.font.color.type == "rgb":
        fg = _normalize_hex(cell.font.color.rgb) or fg
    return fg, bg


def _safe_font_color(background: str, profile: ThemeProfile) -> str:
    candidates = [profile.header_fg, "#000000", "#FFFFFF", profile.accent]
    best = max(candidates, key=lambda color: contrast_ratio(color, background))
    return best if contrast_ratio(best, background) >= MIN_CONTRAST_RATIO else "#000000"


def audit_theme_contrast(workbook_path: str, *, theme: str = "corporate_formal") -> dict:
    workbook = load_workbook(workbook_path)
    profile = get_theme(theme)
    violations = []
    palette = theme_palette(profile)
    off_theme = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fg, bg = _cell_colors(cell, profile)
                ratio = contrast_ratio(fg, bg)
                if ratio < MIN_CONTRAST_RATIO:
                    violations.append({"sheet": ws.title, "cell": cell.coordinate, "foreground": fg, "background": bg, "ratio": round(ratio, 2)})
                for color in (fg, bg):
                    normalized = _normalize_hex(color)
                    if normalized and normalized.upper() not in palette:
                        off_theme.append({"sheet": ws.title, "cell": cell.coordinate, "color": normalized})
    return {
        "contrast_pass": not violations,
        "violations": violations,
        "off_theme": off_theme,
        "off_theme_count": len(off_theme),
    }


def enforce_theme_law(
    workbook_path: str,
    output_path: str | None = None,
    *,
    theme: str = "corporate_formal",
    repair_contrast: bool = True,
) -> dict:
    profile = get_theme(theme)
    path = Path(workbook_path)
    target = Path(output_path) if output_path else path
    workbook = load_workbook(path)
    repairs = []

    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fg, bg = _cell_colors(cell, profile)
                if cell.row == 1 or (cell.fill and cell.fill.fill_type == "solid" and bg != profile.background):
                    if cell.row == 1:
                        bg = profile.header_bg
                        cell.fill = PatternFill("solid", fgColor=profile.header_bg.replace("#", ""))
                    if repair_contrast and contrast_ratio(fg, bg) < MIN_CONTRAST_RATIO:
                        new_fg = _safe_font_color(bg, profile)
                        cell.font = Font(
                            name=cell.font.name,
                            sz=cell.font.sz,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            color=new_fg.replace("#", ""),
                        )
                        repairs.append({"sheet": ws.title, "cell": cell.coordinate, "foreground": new_fg, "background": bg})

    workbook.save(target)
    audit = audit_theme_contrast(str(target), theme=theme)
    return {"output_path": str(target), "repairs": repairs, "audit": audit}
