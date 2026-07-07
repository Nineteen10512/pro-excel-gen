from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class FormulaPlan:
    column: str
    kind: str
    source_column: str | None = None
    numerator_column: str | None = None
    denominator_column: str | None = None
    lookup_value_column: str | None = None
    lookup_table: str | None = None
    return_column: str | None = None
    default: str | int | float | None = None
    formula_mode: str = "modern"


FORMULA_PLAN_FIELDS = set(FormulaPlan.__dataclass_fields__)


def _excel_row(row_zero_based: int, header_row_zero_based: int) -> int:
    return header_row_zero_based + 2 + row_zero_based


def _column_lookup(columns: list[dict]) -> dict[str, int]:
    return {column["header"]: idx for idx, column in enumerate(columns)}


def _cell(header: str, row_idx: int, columns: list[dict]) -> str:
    col_idx = _column_lookup(columns)[header] + 1
    return f"{get_column_letter(col_idx)}{row_idx}"


def downgrade_formula_mode(formula: str, *, formula_mode: str = "compat") -> str:
    """Convert supported modern formulas to conservative legacy equivalents."""

    if formula_mode in {"modern", "auto"}:
        return formula

    match = re.match(r"=XLOOKUP\(([^,]+),([^,]+),([^,]+),([^)]+)\)", formula, re.I)
    if match:
        lookup_value, lookup_range, return_range, default = [item.strip() for item in match.groups()]
        return f'=IFERROR(INDEX({return_range},MATCH({lookup_value},{lookup_range},0)),{default})'
    return formula


def plan_formulas(
    columns: list[dict],
    rows: list[dict],
    formula_plan: list[dict] | dict | None,
    *,
    header_row: int = 0,
    formula_mode: str = "modern",
) -> list[dict]:
    """Return row-level formula assignments for a sheet spec.

    `header_row` is zero-based because xlsxwriter uses zero-based positions.
    """

    if not formula_plan:
        return []
    plans = formula_plan if isinstance(formula_plan, list) else [formula_plan]
    assignments: list[dict] = []
    headers = [column["header"] for column in columns]

    for row_zero_based, _row in enumerate(rows):
        excel_row = _excel_row(row_zero_based, header_row)
        for raw_plan in plans:
            payload = {key: value for key, value in raw_plan.items() if key in FORMULA_PLAN_FIELDS}
            plan = FormulaPlan(**{**payload, "formula_mode": raw_plan.get("formula_mode", formula_mode)})
            if plan.column not in headers:
                continue
            formula = None
            if plan.kind == "percent_of_total" and plan.source_column:
                source_cell = _cell(plan.source_column, excel_row, columns)
                source_col = get_column_letter(_column_lookup(columns)[plan.source_column] + 1)
                first_data = header_row + 2
                last_data = header_row + 1 + len(rows)
                formula = f"=IFERROR({source_cell}/SUM(${source_col}${first_data}:${source_col}${last_data}),0)"
            elif plan.kind == "delta" and plan.numerator_column and plan.denominator_column:
                numerator = _cell(plan.numerator_column, excel_row, columns)
                denominator = _cell(plan.denominator_column, excel_row, columns)
                formula = f"={numerator}-{denominator}"
            elif plan.kind == "ratio" and plan.numerator_column and plan.denominator_column:
                numerator = _cell(plan.numerator_column, excel_row, columns)
                denominator = _cell(plan.denominator_column, excel_row, columns)
                formula = f"=IFERROR({numerator}/{denominator},0)"
            elif plan.kind == "running_total" and plan.source_column:
                source_col = get_column_letter(_column_lookup(columns)[plan.source_column] + 1)
                formula = f"=SUM(${source_col}${header_row + 2}:{source_col}{excel_row})"
            elif plan.kind == "lookup" and plan.lookup_value_column and plan.lookup_table and plan.return_column:
                lookup_value = _cell(plan.lookup_value_column, excel_row, columns)
                default = f'"{plan.default}"' if isinstance(plan.default, str) else (plan.default if plan.default is not None else '""')
                formula = f"=XLOOKUP({lookup_value},{plan.lookup_table}[Key],{plan.lookup_table}[{plan.return_column}],{default})"
                formula = downgrade_formula_mode(formula, formula_mode=plan.formula_mode)

            if formula:
                assignments.append({"row": row_zero_based, "column": plan.column, "formula": formula})
    return assignments


def apply_formula_plan_to_sheet_spec(sheet: dict, *, formula_mode: str = "modern", header_row: int = 0) -> dict:
    """Return a copy of a sheet spec with planned formulas inserted."""

    updated = copy(sheet)
    columns = [dict(column) for column in sheet.get("columns", [])]
    rows = [dict(row) for row in sheet.get("rows", [])]
    formula_plan = sheet.get("formula_plan")
    if not formula_plan:
        return updated

    existing_headers = {column["header"] for column in columns}
    plans = formula_plan if isinstance(formula_plan, list) else [formula_plan]
    for plan in plans:
        column = plan.get("column")
        if column and column not in existing_headers:
            columns.append({"header": column, "format": plan.get("format", "number")})
            existing_headers.add(column)

    assignments = plan_formulas(columns, rows, formula_plan, header_row=header_row, formula_mode=formula_mode)
    for item in assignments:
        target_row = rows[item["row"]]
        target_row[item["column"]] = {"formula": item["formula"], "format": _plan_format(plans, item["column"])}

    updated["columns"] = columns
    updated["rows"] = rows
    return updated


def _plan_format(plans: list[dict], column: str) -> str:
    for plan in plans:
        if plan.get("column") == column:
            return plan.get("format", "number")
    return "number"


def autofill_formulas(
    workbook_path: str,
    output_path: str | None = None,
    *,
    sheet_name: str | None = None,
    table_name: str | None = None,
    start_row: int | None = None,
    end_row: int | None = None,
) -> dict:
    """Fill formulas down from the nearest formula row into blank cells."""

    path = Path(workbook_path)
    target = Path(output_path) if output_path else path
    workbook = load_workbook(path)
    sheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
    filled = 0

    for ws in sheets:
        min_row = start_row or 2
        max_row = end_row or ws.max_row
        if table_name and table_name in ws.tables:
            ref = ws.tables[table_name].ref
            _, bounds = ref.split(":") if ":" in ref else (ref, ref)
            max_row = int(re.sub(r"\D", "", bounds))
        for col_idx in range(1, ws.max_column + 1):
            anchor = None
            anchor_coord = None
            for row_idx in range(min_row, max_row + 1):
                value = ws.cell(row_idx, col_idx).value
                if isinstance(value, str) and value.startswith("="):
                    anchor = value
                    anchor_coord = ws.cell(row_idx, col_idx).coordinate
                    continue
                if anchor and value in (None, ""):
                    dest = ws.cell(row_idx, col_idx).coordinate
                    ws.cell(row_idx, col_idx).value = Translator(anchor, origin=anchor_coord).translate_formula(dest)
                    filled += 1

    workbook.save(target)
    return {"output_path": str(target), "filled_cells": filled}
