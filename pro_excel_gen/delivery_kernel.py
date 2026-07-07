from __future__ import annotations

import csv
import shutil
import subprocess
import tempfile
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from .chart_bridge import build_chart_from_range
from .excel_jsx import generate_from_rows
from .normalizer import normalize_workbook_ui
from .quality import inspect_workbook, scan_formula_issues
from .theme_law import audit_theme_contrast
from .workbook_controls import inspect_visual_layout, create_delivery_audit_sheet


def import_tabular_data(source_path: str, *, sheet_name: str | None = None, limit: int | None = None) -> list[dict]:
    """Import CSV, TSV, or XLSX data as typed row dictionaries."""

    path = Path(source_path)
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle, delimiter=delimiter))
        typed = [_coerce_row(row) for row in rows]
        return typed[:limit] if limit else typed
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=False)
        try:
            ws = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
            headers = [cell.value for cell in ws[1]]
            rows = []
            for raw in ws.iter_rows(min_row=2, values_only=True):
                if all(value is None for value in raw):
                    continue
                rows.append({str(headers[idx] or f"Column {idx + 1}"): value for idx, value in enumerate(raw)})
                if limit and len(rows) >= limit:
                    break
            return rows
        finally:
            workbook.close()
    raise ValueError(f"Unsupported tabular source: {source_path}")


def _coerce_row(row: dict) -> dict:
    return {key: _coerce_value(value) for key, value in row.items()}


def _coerce_value(value):
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return ""
    try:
        if text.endswith("%"):
            return float(text[:-1].replace(",", "")) / 100
        if "." in text:
            return float(text.replace(",", ""))
        return int(text.replace(",", ""))
    except ValueError:
        return value


def analyze_table(rows: list[dict]) -> dict:
    """Return compact row/column and numeric summary metadata."""

    if not rows:
        return {"row_count": 0, "columns": [], "numeric_columns": [], "summaries": {}}
    columns = list(rows[0].keys())
    numeric_columns = []
    summaries = {}
    for column in columns:
        values = [row.get(column) for row in rows if isinstance(row.get(column), (int, float))]
        if not values:
            continue
        numeric_columns.append(column)
        summaries[column] = {
            "count": len(values),
            "sum": sum(values),
            "average": mean(values),
            "min": min(values),
            "max": max(values),
        }
    return {"row_count": len(rows), "columns": columns, "numeric_columns": numeric_columns, "summaries": summaries}


def create_workbook_from_source(
    source_path: str,
    output_path: str,
    *,
    sheet_name: str = "Data",
    theme: str = "corporate_formal",
    add_analysis: bool = True,
    add_chart: bool = True,
) -> dict:
    rows = import_tabular_data(source_path)
    if not rows:
        raise ValueError("Source contains no rows")
    generate_from_rows([{"title": sheet_name, "rows": rows}], output_path, theme=theme)
    if add_analysis:
        add_analysis_sheet(output_path, sheet_name=sheet_name, theme=theme)
    if add_chart:
        analysis = analyze_table(rows)
        category = _first_text_column(rows)
        numeric = analysis["numeric_columns"][:1]
        if category and numeric:
            workbook = load_workbook(output_path)
            ws = workbook[sheet_name]
            headers = [cell.value for cell in ws[1]]
            category_col = headers.index(category) + 1
            value_col = headers.index(numeric[0]) + 1
            workbook.save(output_path)
            build_chart_from_range(
                output_path,
                sheet_name=sheet_name,
                categories_range=f"'{sheet_name}'!${get_column_letter(category_col)}$2:${get_column_letter(category_col)}${ws.max_row}",
                values_ranges=[f"'{sheet_name}'!${get_column_letter(value_col)}$1:${get_column_letter(value_col)}${ws.max_row}"],
                chart_type="column",
                title=f"{numeric[0]} by {category}",
                anchor="H2",
                theme=theme,
            )
    normalize_workbook_ui(output_path, theme=theme)
    verification = verify_delivery_workbook(output_path, render=True)
    return {"output_path": output_path, "analysis": analyze_table(rows), "verification": verification}


def _first_text_column(rows: list[dict]) -> str | None:
    for column in rows[0].keys():
        if any(isinstance(row.get(column), str) and row.get(column) for row in rows):
            return column
    return None


def add_analysis_sheet(workbook_path: str, output_path: str | None = None, *, sheet_name: str | None = None, theme: str = "corporate_formal") -> dict:
    target = output_path or workbook_path
    workbook = load_workbook(workbook_path)
    source = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    if "Analysis" in workbook.sheetnames:
        del workbook["Analysis"]
    ws = workbook.create_sheet("Analysis")
    ws.append(["Column", "Count", "Sum", "Average", "Min", "Max"])
    headers = [cell.value for cell in source[1]]
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        values = [source.cell(row, idx).value for row in range(2, source.max_row + 1)]
        if not any(isinstance(value, (int, float)) for value in values):
            continue
        ws.append(
            [
                header,
                f"=COUNT('{source.title}'!{letter}2:{letter}{source.max_row})",
                f"=SUM('{source.title}'!{letter}2:{letter}{source.max_row})",
                f"=AVERAGE('{source.title}'!{letter}2:{letter}{source.max_row})",
                f"=MIN('{source.title}'!{letter}2:{letter}{source.max_row})",
                f"=MAX('{source.title}'!{letter}2:{letter}{source.max_row})",
            ]
        )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(target)
    normalize_workbook_ui(target, theme=theme)
    return {"output_path": target, "sheet": "Analysis", "formula_rows": max(workbook["Analysis"].max_row - 1, 0)}


def recalculate_workbook(workbook_path: str, output_path: str | None = None) -> dict:
    """Mark formulas for recalculation and use LibreOffice when available."""

    target = output_path or workbook_path
    workbook = load_workbook(workbook_path)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(target)

    office = shutil.which("soffice") or shutil.which("libreoffice")
    if not office:
        return {"output_path": target, "status": "marked_for_recalc", "engine": None}

    with tempfile.TemporaryDirectory(prefix="pro_excel_recalc_") as tmp:
        proc = subprocess.run(
            [office, "--headless", "--convert-to", "xlsx", "--outdir", tmp, target],
            text=True,
            capture_output=True,
        )
    return {"output_path": target, "status": "recalculated" if proc.returncode == 0 else "marked_for_recalc", "engine": "libreoffice", "returncode": proc.returncode}


def render_sheet_preview(
    workbook_path: str,
    output_path: str | None = None,
    *,
    sheet_name: str | None = None,
    max_rows: int = 30,
    max_cols: int = 12,
) -> dict:
    """Render a lightweight visual preview PNG for layout QA."""

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        ws = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = min(ws.max_row, max_rows)
        cols = min(ws.max_column, max_cols)
        col_widths = []
        for col_idx in range(1, cols + 1):
            width = max(90, min(int((ws.column_dimensions[get_column_letter(col_idx)].width or 12) * 8), 260))
            col_widths.append(width)
        row_heights = [30 if row_idx == 1 else 24 for row_idx in range(1, rows + 1)]
        image_width = sum(col_widths) + 1
        image_height = sum(row_heights) + 1
        image = Image.new("RGB", (image_width, image_height), "white")
        draw = ImageDraw.Draw(image)
        font = ImageFont.load_default()
        y = 0
        for row_idx in range(1, rows + 1):
            x = 0
            for col_idx in range(1, cols + 1):
                cell = ws.cell(row_idx, col_idx)
                fill = "#F2F6FA" if row_idx == 1 else "#FFFFFF"
                if cell.fill and cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb:
                    rgb = cell.fill.fgColor.rgb[-6:]
                    fill = f"#{rgb}"
                draw.rectangle([x, y, x + col_widths[col_idx - 1], y + row_heights[row_idx - 1]], fill=fill, outline="#D9E2F3")
                text = "" if cell.value is None else str(cell.value)
                if len(text) > 28:
                    text = text[:25] + "..."
                draw.text((x + 4, y + 6), text, fill="#111111", font=font)
                x += col_widths[col_idx - 1]
            y += row_heights[row_idx - 1]
        out = Path(output_path) if output_path else Path(workbook_path).with_suffix(f".{ws.title}.preview.png")
        image.save(out)
        return {"output_path": str(out), "sheet": ws.title, "width": image_width, "height": image_height}
    finally:
        workbook.close()


def verify_delivery_workbook(workbook_path: str, *, render: bool = True) -> dict:
    metrics = inspect_workbook(workbook_path)
    formula_scan = scan_formula_issues(workbook_path)
    theme_audit = audit_theme_contrast(workbook_path)
    visual_layout = inspect_visual_layout(workbook_path)
    previews = []
    if render:
        workbook = load_workbook(workbook_path, read_only=True)
        try:
            sheet_names = list(workbook.sheetnames)
        finally:
            workbook.close()
        for sheet_name in sheet_names:
            previews.append(render_sheet_preview(workbook_path, sheet_name=sheet_name))
    errors = []
    warnings = []
    if metrics["sheet_count"] == 0:
        errors.append("Workbook has no sheets")
    if formula_scan["error_count"]:
        warnings.append("Formula scan found visible formula error tokens")
    if not theme_audit["contrast_pass"]:
        warnings.append("Theme contrast audit found low-contrast cells")
    if not visual_layout["passed"]:
        warnings.append("Visual layout audit found possible clipping or blank charts")
    return {
        "path": workbook_path,
        "metrics": metrics,
        "formula_scan": formula_scan,
        "theme_audit": theme_audit,
        "visual_layout": visual_layout,
        "previews": previews,
        "errors": errors,
        "warnings": warnings,
        "passed": not errors,
    }


def finalize_delivery_workbook(workbook_path: str, output_path: str, *, add_audit_sheet: bool = True) -> dict:
    """Run final verification, optionally embed audit sheet, and export `.xlsx`."""

    verification = verify_delivery_workbook(workbook_path, render=True)
    working_path = workbook_path
    if add_audit_sheet:
        create_delivery_audit_sheet(workbook_path, report=verification)
        working_path = workbook_path
        verification = verify_delivery_workbook(workbook_path, render=True)
    export = export_xlsx(working_path, output_path, verify=True)
    return {"output_path": output_path, "verification": verification, "export": export}


def export_xlsx(workbook_path: str, output_path: str, *, verify: bool = True) -> dict:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(workbook_path, output_path)
    result = {"output_path": output_path}
    if verify:
        result["verification"] = verify_delivery_workbook(output_path, render=True)
    return result
