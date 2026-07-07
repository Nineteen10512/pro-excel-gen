from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .normalizer import normalize_sheet_ui
from .themes import get_theme


def add_dropdown_validations(
    workbook_path: str,
    output_path: str | None = None,
    *,
    validations: list[dict],
    theme: str = "corporate_formal",
) -> dict:
    """Add editable dropdown validation lists to worksheet ranges."""

    target = output_path or workbook_path
    workbook = load_workbook(workbook_path)
    added = []
    try:
        for item in validations:
            ws = workbook[item["sheet"]]
            options = [str(option).replace('"', '""') for option in item["options"]]
            formula = '"' + ",".join(options) + '"'
            validation = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=item.get("allow_blank", True),
                showDropDown=False,
            )
            validation.error = item.get("error", "Select a value from the list.")
            validation.errorTitle = item.get("error_title", "Invalid value")
            validation.prompt = item.get("prompt", "Choose one allowed value.")
            validation.promptTitle = item.get("prompt_title", item.get("header", "Allowed values"))
            ws.add_data_validation(validation)
            validation.add(item["range"])
            added.append({"sheet": ws.title, "range": item["range"], "options": options})
        workbook.save(target)
    finally:
        workbook.close()
    normalize_sheet_ui(target, theme=theme)
    return {"output_path": target, "validations": added}


def apply_semantic_conditional_formats(
    workbook_path: str,
    output_path: str | None = None,
    *,
    rules: list[dict],
    theme: str = "corporate_formal",
) -> dict:
    """Apply semantic conditional formatting using theme colors."""

    profile = get_theme(theme)
    target = output_path or workbook_path
    workbook = load_workbook(workbook_path)
    applied = []
    try:
        for item in rules:
            ws = workbook[item["sheet"]]
            cell_range = item["range"]
            rule_type = item.get("type", "color_scale")
            if rule_type == "color_scale":
                ws.conditional_formatting.add(
                    cell_range,
                    ColorScaleRule(
                        start_type="min",
                        start_color=profile.bad.replace("#", ""),
                        mid_type="percentile",
                        mid_value=50,
                        mid_color=profile.warn.replace("#", ""),
                        end_type="max",
                        end_color=profile.good.replace("#", ""),
                    ),
                )
            elif rule_type == "greater_than":
                fill = PatternFill("solid", fgColor=profile.good.replace("#", ""))
                ws.conditional_formatting.add(cell_range, CellIsRule(operator="greaterThan", formula=[str(item["value"])], fill=fill))
            elif rule_type == "less_than":
                fill = PatternFill("solid", fgColor=profile.bad.replace("#", ""))
                ws.conditional_formatting.add(cell_range, CellIsRule(operator="lessThan", formula=[str(item["value"])], fill=fill))
            elif rule_type == "formula":
                fill = PatternFill("solid", fgColor=item.get("fill", profile.warn).replace("#", ""))
                ws.conditional_formatting.add(cell_range, FormulaRule(formula=[item["formula"]], fill=fill))
            else:
                continue
            applied.append({"sheet": ws.title, "range": cell_range, "type": rule_type})
        workbook.save(target)
    finally:
        workbook.close()
    return {"output_path": target, "rules": applied}


def inspect_visual_layout(workbook_path: str, *, max_text_to_width_ratio: float = 1.35) -> dict:
    """Detect common visual defects before delivery."""

    workbook = load_workbook(workbook_path, data_only=False)
    issues = []
    try:
        for ws in workbook.worksheets:
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                width = ws.column_dimensions[letter].width or 8.43
                max_len = 0
                max_cell = None
                for row_idx in range(1, min(ws.max_row, 200) + 1):
                    value = ws.cell(row_idx, col_idx).value
                    length = len(str(value)) if value is not None else 0
                    if length > max_len:
                        max_len = length
                        max_cell = ws.cell(row_idx, col_idx).coordinate
                if width and max_len / width > max_text_to_width_ratio and max_len > 18:
                    issues.append(
                        {
                            "sheet": ws.title,
                            "cell": max_cell,
                            "issue": "possible_clipping",
                            "width": round(width, 2),
                            "text_len": max_len,
                        }
                    )
            for chart_idx, chart in enumerate(getattr(ws, "_charts", []), start=1):
                if not getattr(chart, "series", None):
                    issues.append({"sheet": ws.title, "chart_index": chart_idx, "issue": "blank_chart"})
    finally:
        workbook.close()
    return {"passed": not issues, "issues": issues, "issue_count": len(issues)}


def create_delivery_audit_sheet(workbook_path: str, output_path: str | None = None, *, report: dict, sheet_name: str = "DeliveryAudit") -> dict:
    target = output_path or workbook_path
    workbook = load_workbook(workbook_path)
    try:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)
        ws.append(["Section", "Key", "Value"])
        for section, values in report.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    ws.append([section, key, str(value)[:500]])
            else:
                ws.append([section, "", str(values)[:500]])
        workbook.save(target)
    finally:
        workbook.close()
    normalize_sheet_ui(target, sheet_name=sheet_name)
    return {"output_path": target, "sheet": sheet_name}
