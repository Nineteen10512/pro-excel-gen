from __future__ import annotations

import tempfile
import traceback
import zipfile
from pathlib import Path
import sys

from PIL import Image
from openpyxl import load_workbook


BUNDLE_ROOT = Path(__file__).resolve().parents[1]
if str(BUNDLE_ROOT) not in sys.path:
    sys.path.insert(0, str(BUNDLE_ROOT))


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _write_sample_png(path: Path) -> None:
    Image.new("RGB", (160, 90), color=(30, 110, 180)).save(path)


def _count_zip_prefix(path: Path, prefix: str) -> int:
    with zipfile.ZipFile(path) as zf:
        return sum(1 for name in zf.namelist() if name.startswith(prefix))


def _build_sample_workbook(out_path: Path, image_path: Path) -> None:
    import pro_excel_gen

    spec = {
        "meta": {"title": "Image Sample"},
        "sheets": [
            {
                "title": "概览",
                "freeze_panes": [1, 0],
                "columns": [
                    {"header": "指标", "width": 16},
                    {"header": "说明", "width": 18},
                    {"header": "数值", "format": "number"},
                    {"header": "备注", "width": 18},
                ],
                "rows": [
                    {"指标": "收入", "说明": "核心增长指标", "数值": 1280, "备注": "图片右侧"},
                    {"指标": "毛利率", "说明": "翻译后可能更长", "数值": 0.42, "备注": "保持格式"},
                    {"指标": "公式合计", "说明": "不要翻译公式", "数值": {"formula": "=SUM(C2:C3)", "format": "number"}, "备注": "公式保留"},
                ],
                "table": {"name": "ImageSampleTable", "style": "Table Style Medium 4"},
                "images": [{"path": str(image_path), "cell": "F2", "x_scale": 0.7, "y_scale": 0.7}],
            }
        ],
    }
    pro_excel_gen.generate(spec, str(out_path))


def test_image_preservation(tmp: Path) -> None:
    import pro_excel_gen

    image_path = tmp / "sample.png"
    src = tmp / "image_sample.xlsx"
    out = tmp / "image_sample.en.xlsx"
    _write_sample_png(image_path)
    _build_sample_workbook(src, image_path)

    segments = pro_excel_gen.collect_translation_segments(str(src))
    translations = []
    for item in segments:
        text = item["text"]
        if not text.strip():
            continue
        translations.append({"id": item["id"], "translation": f"EN {text}"})
    pro_excel_gen.apply_translation_map(str(src), translations, output_path=str(out), target_lang="en")

    _assert(_count_zip_prefix(src, "xl/media/") == _count_zip_prefix(out, "xl/media/"), "Media count changed")
    _assert(_count_zip_prefix(src, "xl/drawings/") == _count_zip_prefix(out, "xl/drawings/"), "Drawing count changed")

    workbook = load_workbook(out, data_only=False)
    ws = workbook["EN 概览"]
    _assert(ws["A2"].value == "EN 收入", "Translated cell text missing")
    _assert(isinstance(ws["C4"].value, str) and ws["C4"].value.startswith("="), "Formula cell lost formula")


def run_suite(tmp: Path) -> int:
    failures = 0
    try:
        test_image_preservation(tmp)
        print("PASS XLSX image preservation")
    except Exception:
        failures += 1
        print("FAIL XLSX image preservation")
        traceback.print_exc()
    return failures


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="pro_excel_image_regression_") as tmp_dir:
        return run_suite(Path(tmp_dir))


if __name__ == "__main__":
    raise SystemExit(main())
