from __future__ import annotations

import argparse
import shutil
import sys
import tempfile
import traceback
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from run_image_regression import run_suite as run_image_regression_suite


BUNDLE_ROOT = Path(__file__).resolve().parents[1]
if str(BUNDLE_ROOT) not in sys.path:
    sys.path.insert(0, str(BUNDLE_ROOT))


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _read_frontmatter(path: Path) -> dict[str, str]:
    text = path.read_text(encoding="utf-8")
    _assert(text.startswith("---\n"), f"{path} missing YAML frontmatter")
    raw = text.split("---", 2)[1]
    data: dict[str, str] = {}
    for line in raw.splitlines():
        if ":" in line:
            key, value = line.split(":", 1)
            data[key.strip()] = value.strip()
    return data


def _assert_xlsx(path: Path) -> None:
    _assert(path.exists(), f"{path.name} not created")
    _assert(path.stat().st_size > 2000, f"{path.name} too small")
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
    _assert("xl/workbook.xml" in names, f"{path.name} missing xl/workbook.xml")


def test_skill_frontmatter() -> None:
    meta = _read_frontmatter(BUNDLE_ROOT / "SKILL.md")
    _assert(meta.get("name") == "pro-excel-gen", "Skill name mismatch")
    desc = meta.get("description", "")
    _assert(desc.startswith("Use when "), "Skill description must start with 'Use when '")
    _assert(len(desc) < 1024, "Skill description too long")


def test_public_imports() -> None:
    import pro_excel_gen

    for name in (
        "generate",
        "generate_from_rows",
        "build_kpi_dashboard",
        "build_chart_sheet",
        "render_chart_to_png",
        "collect_translation_segments",
        "apply_translation_map",
        "plan_formulas",
        "autofill_formulas",
        "normalize_workbook_ui",
        "enforce_theme_law",
        "audit_theme_contrast",
        "build_chart_from_table",
        "extract_chart_source_data",
        "export_chart_data_to_sheet",
        "edit_existing_workbook",
        "import_tabular_data",
        "analyze_table",
        "create_workbook_from_source",
        "add_analysis_sheet",
        "recalculate_workbook",
        "render_sheet_preview",
        "verify_delivery_workbook",
        "export_xlsx",
        "scan_formula_issues",
        "finalize_delivery_workbook",
        "add_dropdown_validations",
        "apply_semantic_conditional_formats",
        "inspect_visual_layout",
        "create_delivery_audit_sheet",
    ):
        _assert(hasattr(pro_excel_gen, name), f"Missing pro_excel_gen.{name}")


def test_generation(out_dir: Path) -> None:
    import pro_excel_gen

    spec = pro_excel_gen.build_kpi_dashboard(
        "Smoke Dashboard",
        metrics=[
            {"metric": "Revenue", "value": 1200000, "format": "currency", "owner": "CFO", "status": "On track"},
            {"metric": "Margin", "value": 0.41, "format": "percent", "owner": "FP&A", "status": "At risk"},
        ],
        series={"months": ["Jan", "Feb", "Mar"], "series": [{"name": "ARR", "values": [120, 130, 145]}]},
    )
    output = out_dir / "smoke_dashboard.xlsx"
    pro_excel_gen.generate(spec, str(output), theme="corporate_formal")
    _assert_xlsx(output)
    workbook = load_workbook(output, data_only=False)
    _assert("Dashboard" in workbook.sheetnames, "Dashboard sheet missing")
    _assert("Trend" in workbook.sheetnames, "Trend sheet missing")
    _assert(workbook["Dashboard"]["B4"].data_type in {"n", "f"}, "Numeric cell wrong type")


def test_generate_from_rows(out_dir: Path) -> None:
    import pro_excel_gen

    output = out_dir / "smoke_rows.xlsx"
    pro_excel_gen.generate_from_rows(
        [{"title": "Rows", "rows": [{"Name": "Ada", "Score": 98}, {"Name": "Grace", "Score": 99}]}],
        str(output),
    )
    _assert_xlsx(output)


def test_formula_first_generation(out_dir: Path) -> None:
    import pro_excel_gen

    output = out_dir / "formula_first.xlsx"
    spec = {
        "meta": {"formula_mode": "modern"},
        "sheets": [
            {
                "title": "FormulaFirst",
                "columns": [
                    {"header": "Product"},
                    {"header": "Revenue", "format": "currency"},
                    {"header": "Cost", "format": "currency"},
                ],
                "rows": [
                    {"Product": "Alpha", "Revenue": 100, "Cost": 60},
                    {"Product": "Beta", "Revenue": 300, "Cost": 120},
                ],
                "formula_plan": [
                    {"column": "Gross Profit", "kind": "delta", "numerator_column": "Revenue", "denominator_column": "Cost", "format": "currency"},
                    {"column": "Revenue Share", "kind": "percent_of_total", "source_column": "Revenue", "format": "percent"},
                ],
                "table": {"name": "FormulaTable"},
            }
        ],
    }
    pro_excel_gen.generate(spec, str(output))
    workbook = load_workbook(output, data_only=False)
    ws = workbook["FormulaFirst"]
    _assert(ws["D2"].value == "=B2-C2", "Delta formula not generated")
    _assert(str(ws["E2"].value).startswith("=IFERROR(B2/SUM($B$2:$B$3),0)"), "Percent formula not generated")


def test_theme_law_and_contrast(out_dir: Path) -> None:
    import pro_excel_gen
    from openpyxl.styles import Font, PatternFill

    output = out_dir / "theme_law.xlsx"
    pro_excel_gen.generate_from_rows([{"title": "Theme", "rows": [{"Name": "Alpha", "Value": 1}]}], str(output))
    workbook = load_workbook(output)
    ws = workbook["Theme"]
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["A1"].font = Font(color="FFFFFF", bold=True)
    workbook.save(output)

    before = pro_excel_gen.audit_theme_contrast(str(output))
    _assert(not before["contrast_pass"], "Low contrast fixture did not fail audit")
    result = pro_excel_gen.enforce_theme_law(str(output), theme="corporate_formal")
    _assert(result["audit"]["contrast_pass"], "Theme law did not repair contrast")


def test_chart_bridge_and_existing_edit(out_dir: Path) -> None:
    import pro_excel_gen

    output = out_dir / "chart_bridge.xlsx"
    pro_excel_gen.generate_from_rows(
        [{"title": "Rows", "rows": [{"Name": "Ada", "Score": 98}, {"Name": "Grace", "Score": 99}, {"Name": "Lin", "Score": 94}]}],
        str(output),
    )
    insert = pro_excel_gen.insert_chart_into_table_region(
        str(output),
        sheet_name="Rows",
        table_name="RowsTable",
        categories_column="Name",
        series_columns=["Score"],
        title="Scores",
        anchor="E2",
    )
    _assert(insert["series_count"] == 1, "Table chart insertion failed")
    extracted = pro_excel_gen.extract_chart_source_data(str(output))
    _assert(extracted and extracted[0]["series"], "Chart source extraction returned no series")
    _assert(extracted[0]["series"][0]["values"] == [98, 99, 94], "Chart values were not extracted")
    export = pro_excel_gen.export_chart_data_to_sheet(str(output), sheet_name="ExtractedChartData")
    _assert(export["rows"] == 3, "Chart data export row count mismatch")

    edit_output = out_dir / "chart_bridge_edited.xlsx"
    audit = pro_excel_gen.edit_existing_workbook(
        str(output),
        str(edit_output),
        [
            {"type": "normalize_sheet_ui", "sheet": "Rows"},
            {"type": "embed_microchart_column", "sheet": "Rows", "source_column": 2, "target_header": "Score Bar"},
        ],
    )
    _assert_xlsx(edit_output)
    edited = load_workbook(edit_output, data_only=False)
    _assert(str(edited["Rows"]["C2"].value).startswith("=REPT"), "Microchart formula not inserted")
    _assert(len(audit["operations"]) == 2, "Edit audit missing operations")


def test_image_chart_inference_stub() -> None:
    import pro_excel_gen

    result = pro_excel_gen.infer_chart_data_from_image("chart.png")
    _assert(result["review_required"] is True, "Image chart inference must require review")
    _assert(result["confidence"] == 0.0, "Image chart inference confidence should be explicit")


def test_delivery_kernel_csv_analysis_visual_export(out_dir: Path) -> None:
    import pro_excel_gen

    csv_path = out_dir / "source.csv"
    csv_path.write_text("Month,Revenue,Cost\nJan,100,60\nFeb,140,70\nMar,180,90\n", encoding="utf-8")
    rows = pro_excel_gen.import_tabular_data(str(csv_path))
    analysis = pro_excel_gen.analyze_table(rows)
    _assert(analysis["row_count"] == 3, "CSV import row count mismatch")
    _assert(analysis["summaries"]["Revenue"]["sum"] == 420, "Analysis summary mismatch")

    workbook_path = out_dir / "from_csv.xlsx"
    result = pro_excel_gen.create_workbook_from_source(str(csv_path), str(workbook_path))
    _assert_xlsx(workbook_path)
    _assert(result["verification"]["passed"], "Delivery verification failed")
    workbook = load_workbook(workbook_path, data_only=False)
    _assert("Analysis" in workbook.sheetnames, "Analysis sheet missing")
    _assert(str(workbook["Analysis"]["C2"].value).startswith("=SUM("), "Analysis formula missing")

    preview = pro_excel_gen.render_sheet_preview(str(workbook_path), sheet_name="Data")
    _assert(Path(preview["output_path"]).exists(), "Preview PNG missing")
    formula_scan = pro_excel_gen.scan_formula_issues(str(workbook_path))
    _assert(formula_scan["error_count"] == 0, "Unexpected formula issue")
    recalc = pro_excel_gen.recalculate_workbook(str(workbook_path))
    _assert(recalc["status"] in {"marked_for_recalc", "recalculated"}, "Recalc status invalid")
    exported = out_dir / "exported.xlsx"
    export = pro_excel_gen.export_xlsx(str(workbook_path), str(exported))
    _assert_xlsx(exported)
    _assert(export["verification"]["passed"], "Export verification failed")


def test_controls_visual_audit_and_finalize(out_dir: Path) -> None:
    import pro_excel_gen

    workbook_path = out_dir / "controls.xlsx"
    pro_excel_gen.generate_from_rows(
        [{"title": "Controls", "rows": [{"Task": "A", "Status": "Open", "Score": 80}, {"Task": "B", "Status": "Closed", "Score": 55}]}],
        str(workbook_path),
    )
    pro_excel_gen.add_dropdown_validations(
        str(workbook_path),
        validations=[{"sheet": "Controls", "range": "B2:B20", "options": ["Open", "Closed", "At risk"], "header": "Status"}],
    )
    pro_excel_gen.apply_semantic_conditional_formats(
        str(workbook_path),
        rules=[{"sheet": "Controls", "range": "C2:C20", "type": "color_scale"}],
    )
    workbook = load_workbook(workbook_path)
    _assert(workbook["Controls"].data_validations.count >= 1, "Data validation not added")
    workbook.close()

    visual = pro_excel_gen.inspect_visual_layout(str(workbook_path))
    _assert("issues" in visual, "Visual layout audit missing issues field")
    verification = pro_excel_gen.verify_delivery_workbook(str(workbook_path))
    pro_excel_gen.create_delivery_audit_sheet(str(workbook_path), report=verification)
    audited = load_workbook(workbook_path)
    _assert("DeliveryAudit" in audited.sheetnames, "Delivery audit sheet missing")
    audited.close()

    final_path = out_dir / "controls_final.xlsx"
    final = pro_excel_gen.finalize_delivery_workbook(str(workbook_path), str(final_path))
    _assert_xlsx(final_path)
    _assert(final["export"]["verification"]["passed"], "Final delivery export did not pass")


def test_chart_png(out_dir: Path) -> None:
    import pro_excel_gen

    output = out_dir / "chart.png"
    pro_excel_gen.render_chart_to_png(
        {
            "type": "column",
            "title": "Quarterly Revenue",
            "categories": ["Q1", "Q2", "Q3"],
            "series": [{"name": "2026", "values": [12, 18, 25]}],
        },
        str(output),
    )
    _assert(output.exists() and output.stat().st_size > 1000, "Chart PNG missing")
    _assert(output.read_bytes()[:8] == b"\x89PNG\r\n\x1a\n", "Chart output not PNG")


def test_translation_flow(out_dir: Path) -> None:
    import pro_excel_gen

    source = out_dir / "translate_source.xlsx"
    image_path = out_dir / "logo.png"
    from PIL import Image

    Image.new("RGB", (80, 40), color=(200, 60, 40)).save(image_path)
    spec = {
        "meta": {"title": "Translate"},
        "sheets": [
            {
                "title": "中文表",
                "columns": [
                    {"header": "项目", "width": 14},
                    {"header": "说明", "width": 18},
                    {"header": "数值", "format": "number"},
                ],
                "rows": [
                    {"项目": "收入", "说明": "本期增长很快", "数值": 1200},
                    {"项目": "合计", "说明": "保留公式", "数值": {"formula": "=SUM(C2:C2)", "format": "number"}},
                ],
                "table": {"name": "CnTable", "style": "Table Style Medium 3"},
                "images": [{"path": str(image_path), "cell": "E2", "x_scale": 0.8, "y_scale": 0.8}],
            }
        ],
    }
    pro_excel_gen.generate(spec, str(source))
    segments = pro_excel_gen.collect_translation_segments(str(source))
    prompt = pro_excel_gen.build_translation_prompt(segments, source_lang="zh", target_lang="en")
    _assert("Do not add any fact" in prompt, "Prompt missing hard rule")
    translations = [{"id": item["id"], "translation": f"EN {item['text']}"} for item in segments if item["text"].strip()]
    risk = pro_excel_gen.assess_translation_risk(str(source), translations)
    _assert("sheets" in risk, "Risk report missing sheets")
    output = out_dir / "translate_output.xlsx"
    pro_excel_gen.apply_translation_map(str(source), translations, output_path=str(output), target_lang="en")
    _assert_xlsx(output)
    workbook = load_workbook(output, data_only=False)
    _assert("EN 中文表" in workbook.sheetnames, "Translated sheet name missing")
    _assert(workbook["EN 中文表"]["A2"].value == "EN 收入", "Translated cell missing")
    _assert(isinstance(workbook["EN 中文表"]["C3"].value, str) and workbook["EN 中文表"]["C3"].value.startswith("="), "Formula broken")


def test_quality_gate(out_dir: Path) -> None:
    from quality_gates.run_quality_gate import inspect_file

    targets = [out_dir / "smoke_dashboard.xlsx", out_dir / "smoke_rows.xlsx", out_dir / "translate_output.xlsx"]
    for target in targets:
        report = inspect_file(target)
        _assert(not report["errors"], f"Quality gate errors for {target.name}: {report['errors']}")


def test_templates() -> None:
    import pro_excel_gen

    templates = pro_excel_gen.list_templates()
    _assert(len(templates) >= 5, "Expected template registry")
    recommended = pro_excel_gen.recommend_template("Need sales pipeline and revenue tracker")
    _assert(recommended["name"] == "sales_tracker", "Template recommendation mismatch")


def test_image_regression(out_dir: Path) -> None:
    regression_dir = out_dir / "image_regression"
    if regression_dir.exists():
        shutil.rmtree(regression_dir)
    regression_dir.mkdir(parents=True)
    failures = run_image_regression_suite(regression_dir)
    _assert(failures == 0, "Image regression suite reported failures")


def test_data_cleaner() -> None:
    import pro_excel_gen

    rows = [
        {"产品": " A ", "销量": 100, "单价": " 9.9 "},
        {"产品": "B", "销量": None, "单价": 19.9},
        {"产品": " A ", "销量": 100, "单价": " 9.9 "},
        {"产品": "C", "销量": 9999, "单价": 14.9},
    ]
    cleaned = pro_excel_gen.clean_whitespace(rows)
    _assert(cleaned[0]["产品"] == "A", "Whitespace not cleaned")
    filled = pro_excel_gen.clean_missing(rows, strategy="fill_median")
    _assert(filled[1]["销量"] is not None, "Missing not filled")
    deduped = pro_excel_gen.clean_duplicates(rows, subset=["产品"])
    _assert(len(deduped) < len(rows), "Duplicates not removed")
    typed = pro_excel_gen.type_infer_and_cast(rows)
    _assert(isinstance(typed[0]["单价"], (int, float)), "Type not inferred")


def test_statistics() -> None:
    import pro_excel_gen

    rows = [
        {"月份": "1月", "销售额": 100, "成本": 60},
        {"月份": "2月", "销售额": 120, "成本": 70},
        {"月份": "3月", "销售额": 140, "成本": 80},
        {"月份": "4月", "销售额": 160, "成本": 90},
    ]
    stats = pro_excel_gen.descriptive_stats(rows)
    _assert("销售额" in stats, "Descriptive stats missing")
    _assert("mean" in stats["销售额"], "Stats missing mean")
    corr = pro_excel_gen.correlation_analysis(rows)
    _assert(len(corr["pairs"]) > 0, "Correlation empty")
    trend = pro_excel_gen.trend_analysis(rows, date_column="月份", value_column="销售额")
    _assert(trend["trend_direction"] in ("up", "down", "flat"), "Trend invalid")


def test_ml_tools() -> None:
    import pro_excel_gen

    rows = [
        {"f1": 1.0, "f2": 2.0, "c": "A"}, {"f1": 2.0, "f2": 1.0, "c": "B"},
        {"f1": 3.0, "f2": 3.0, "c": "A"}, {"f1": 4.0, "f2": 2.0, "c": "B"},
        {"f1": 1.5, "f2": 2.5, "c": "A"}, {"f1": 2.5, "f2": 1.5, "c": "B"},
        {"f1": 3.5, "f2": 3.5, "c": "A"}, {"f1": 4.5, "f2": 2.5, "c": "B"},
        {"f1": 1.2, "f2": 2.2, "c": "A"}, {"f1": 2.2, "f2": 1.2, "c": "B"},
    ]
    result = pro_excel_gen.auto_classify(rows, target_column="c")
    _assert("accuracy" in result, "Classify missing accuracy")
    rows_reg = [
        {"x1": 1.0, "x2": 2.0, "y": 3.0}, {"x1": 2.0, "x2": 1.0, "y": 4.0},
        {"x1": 3.0, "x2": 3.0, "y": 6.0}, {"x1": 4.0, "x2": 2.0, "y": 7.0},
        {"x1": 1.5, "x2": 2.5, "y": 3.5}, {"x1": 2.5, "x2": 1.5, "y": 4.5},
        {"x1": 3.5, "x2": 3.5, "y": 6.5}, {"x1": 4.5, "x2": 2.5, "y": 7.5},
        {"x1": 1.2, "x2": 2.2, "y": 3.2}, {"x1": 2.2, "x2": 1.2, "y": 4.2},
    ]
    reg = pro_excel_gen.auto_regress(rows_reg, target_column="y")
    _assert("r2" in reg, "Regress missing r2")
    fi = pro_excel_gen.feature_importance(rows_reg, target_column="y")
    _assert(len(fi) == 2, "Feature importance count")


def run(out_dir: Path) -> int:
    tests = [
        ("skill frontmatter", test_skill_frontmatter),
        ("public imports", test_public_imports),
        ("generation", lambda: test_generation(out_dir)),
        ("generate_from_rows", lambda: test_generate_from_rows(out_dir)),
        ("formula-first generation", lambda: test_formula_first_generation(out_dir)),
        ("theme law and contrast", lambda: test_theme_law_and_contrast(out_dir)),
        ("chart bridge and existing edit", lambda: test_chart_bridge_and_existing_edit(out_dir)),
        ("image chart inference stub", test_image_chart_inference_stub),
        ("delivery kernel csv analysis visual export", lambda: test_delivery_kernel_csv_analysis_visual_export(out_dir)),
        ("controls visual audit and finalize", lambda: test_controls_visual_audit_and_finalize(out_dir)),
        ("chart PNG", lambda: test_chart_png(out_dir)),
        ("translation flow", lambda: test_translation_flow(out_dir)),
        ("templates", test_templates),
        ("image regression", lambda: test_image_regression(out_dir)),
        ("quality gate", lambda: test_quality_gate(out_dir)),
        ("data cleaner", test_data_cleaner),
        ("statistics", test_statistics),
        ("ml tools", test_ml_tools),
    ]
    failures = 0
    for name, fn in tests:
        try:
            fn()
            print(f"PASS {name}")
        except Exception:
            failures += 1
            print(f"FAIL {name}")
            traceback.print_exc()
    return failures


def main() -> int:
    parser = argparse.ArgumentParser(description="Run PRO-EXCEL smoke tests.")
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()

    if args.output_dir:
        out_dir = args.output_dir.resolve()
        if out_dir.exists():
            shutil.rmtree(out_dir)
        out_dir.mkdir(parents=True)
        return run(out_dir)

    with tempfile.TemporaryDirectory(prefix="pro_excel_smoke_") as tmp_dir:
        return run(Path(tmp_dir))


if __name__ == "__main__":
    raise SystemExit(main())
