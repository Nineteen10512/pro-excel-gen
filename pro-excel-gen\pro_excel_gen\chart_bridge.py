from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.cell import range_boundaries

from .normalizer import normalize_sheet_ui


def _chart_class(chart_type: str):
    if chart_type == "line":
        return LineChart
    if chart_type == "pie":
        return PieChart
    return BarChart


def _split_range(range_ref: str) -> tuple[str | None, str]:
    if "!" not in range_ref:
        return None, range_ref.replace("$", "")
    sheet, cells = range_ref.split("!", 1)
    return sheet.strip("'"), cells.replace("$", "")


def _make_reference(workbook, default_sheet: str, range_ref: str) -> Reference:
    sheet_name, cells = _split_range(range_ref)
    ws = workbook[sheet_name or default_sheet]
    min_col, min_row, max_col, max_row = range_boundaries(cells)
    return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def build_chart_from_range(
    workbook_path: str,
    output_path: str | None = None,
    *,
    sheet_name: str,
    categories_range: str,
    values_ranges: list[str],
    chart_type: str = "column",
    title: str = "Chart",
    anchor: str = "H2",
    theme: str = "corporate_formal",
) -> dict:
    path = Path(workbook_path)
    target = Path(output_path) if output_path else path
    workbook = load_workbook(path)
    ws = workbook[sheet_name]
    chart = _chart_class(chart_type)()
    chart.title = title
    chart.style = 10
    categories = _make_reference(workbook, sheet_name, categories_range)
    for value_range in values_ranges:
        data = _make_reference(workbook, sheet_name, value_range)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)
    workbook.save(target)
    normalize_sheet_ui(str(target), sheet_name=sheet_name, theme=theme)
    return {"output_path": str(target), "sheet": sheet_name, "anchor": anchor, "series_count": len(values_ranges)}


def build_chart_from_table(
    workbook_path: str,
    output_path: str | None = None,
    *,
    sheet_name: str,
    table_name: str,
    categories_column: str,
    series_columns: list[str],
    chart_type: str = "column",
    title: str = "Chart",
    anchor: str = "H2",
    theme: str = "corporate_formal",
) -> dict:
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [ws.cell(min_row, col_idx).value for col_idx in range(min_col, max_col + 1)]
    header_to_col = {header: min_col + idx for idx, header in enumerate(headers)}
    cat_col = header_to_col[categories_column]
    values_ranges = [f"'{sheet_name}'!${ws.cell(min_row, header_to_col[name]).column_letter}${min_row}:${ws.cell(max_row, header_to_col[name]).column_letter}${max_row}" for name in series_columns]
    categories_range = f"'{sheet_name}'!${ws.cell(min_row + 1, cat_col).column_letter}${min_row + 1}:${ws.cell(max_row, cat_col).column_letter}${max_row}"
    return build_chart_from_range(
        workbook_path,
        output_path=output_path,
        sheet_name=sheet_name,
        categories_range=categories_range,
        values_ranges=values_ranges,
        chart_type=chart_type,
        title=title,
        anchor=anchor,
        theme=theme,
    )


def insert_chart_into_table_region(workbook_path: str, output_path: str | None = None, **kwargs) -> dict:
    return build_chart_from_table(workbook_path, output_path=output_path, **kwargs)


def _range_values(workbook, range_ref: str) -> list:
    sheet_name, cells = _split_range(range_ref)
    ws = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cells)
    values = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        if len(row) == 1:
            values.append(row[0])
        else:
            values.append(list(row))
    return values


def extract_chart_source_data(workbook_path: str) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=False)
    results = []
    for ws in workbook.worksheets:
        for chart_idx, chart in enumerate(getattr(ws, "_charts", []), start=1):
            chart_item = {"sheet": ws.title, "chart_index": chart_idx, "title": str(chart.title) if chart.title else "", "series": []}
            for series in chart.series:
                values_ref = getattr(getattr(series, "val", None), "numRef", None)
                categories_ref = getattr(getattr(series, "cat", None), "strRef", None) or getattr(getattr(series, "cat", None), "numRef", None)
                value_formula = getattr(values_ref, "f", None)
                category_formula = getattr(categories_ref, "f", None)
                if not value_formula:
                    continue
                chart_item["series"].append(
                    {
                        "name": str(getattr(series, "tx", "")),
                        "values_range": value_formula,
                        "categories_range": category_formula,
                        "values": _range_values(workbook, value_formula),
                        "categories": _range_values(workbook, category_formula) if category_formula else [],
                    }
                )
            results.append(chart_item)
    return results


def export_chart_data_to_sheet(workbook_path: str, output_path: str | None = None, *, sheet_name: str = "ChartData") -> dict:
    target = Path(output_path) if output_path else Path(workbook_path)
    workbook = load_workbook(workbook_path)
    data = extract_chart_source_data(workbook_path)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    ws.append(["Chart Sheet", "Chart Index", "Category", "Series", "Value", "Values Range"])
    rows = 0
    for chart in data:
        for series_idx, series in enumerate(chart["series"], start=1):
            categories = series.get("categories") or list(range(1, len(series.get("values", [])) + 1))
            for category, value in zip(categories, series.get("values", [])):
                ws.append([chart["sheet"], chart["chart_index"], category, series_idx, value, series["values_range"]])
                rows += 1
    workbook.save(target)
    normalize_sheet_ui(str(target), sheet_name=sheet_name)
    return {"output_path": str(target), "sheet": sheet_name, "rows": rows, "charts": len(data)}


def infer_chart_data_from_image(image_path: str) -> dict:
    """Auxiliary placeholder for image-level chart reconstruction."""

    return {
        "image_path": str(image_path),
        "confidence": 0.0,
        "data": [],
        "review_required": True,
        "note": "Image-level chart reconstruction requires visual/model extraction; workbook-native chart extraction is preferred.",
    }


def embed_microchart_column(
    workbook_path: str,
    output_path: str | None = None,
    *,
    sheet_name: str,
    source_column: int,
    target_header: str = "Trend",
    start_row: int = 2,
    end_row: int | None = None,
) -> dict:
    target = Path(output_path) if output_path else Path(workbook_path)
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    end = end_row or ws.max_row
    target_col = ws.max_column + 1
    ws.cell(1, target_col).value = target_header
    source_letter = ws.cell(1, source_column).column_letter
    for row_idx in range(start_row, end + 1):
        ws.cell(row_idx, target_col).value = f'=REPT("|",ROUND({source_letter}{row_idx}/MAX(${source_letter}${start_row}:${source_letter}${end})*10,0))'
    workbook.save(target)
    normalize_sheet_ui(str(target), sheet_name=sheet_name)
    return {"output_path": str(target), "sheet": sheet_name, "column": target_col, "rows": max(end - start_row + 1, 0)}
