from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .theme_law import enforce_theme_law
from .themes import get_theme


def _display_len(value) -> int:
    if value is None:
        return 0
    return len(str(value))


def normalize_sheet_ui(workbook_path: str, output_path: str | None = None, *, sheet_name: str | None = None, theme: str = "corporate_formal") -> dict:
    path = Path(workbook_path)
    target = Path(output_path) if output_path else path
    workbook = load_workbook(path)
    profile = get_theme(theme)
    sheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
    thin = Side(style="thin", color=profile.grid.replace("#", ""))
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    changed_sheets = []

    for ws in sheets:
        changed_sheets.append(ws.title)
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=profile.header_bg.replace("#", ""))
                cell.font = Font(bold=True, color=profile.header_fg.replace("#", ""))
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=_display_len(cell.value) > 24)
                if isinstance(cell.value, float):
                    cell.number_format = "0.0%" if abs(cell.value) <= 1 else "#,##0.0"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"

        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row_idx in range(1, ws.max_row + 1):
                max_len = max(max_len, _display_len(ws.cell(row_idx, col_idx).value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 36)

        for row_idx in range(1, ws.max_row + 1):
            max_len = max(_display_len(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1))
            ws.row_dimensions[row_idx].height = 28 if max_len > 40 else 18

    workbook.save(target)
    law = enforce_theme_law(str(target), theme=theme)
    return {"output_path": str(target), "sheets": changed_sheets, "theme_law": law}


def normalize_workbook_ui(workbook_path: str, output_path: str | None = None, *, theme: str = "corporate_formal") -> dict:
    return normalize_sheet_ui(workbook_path, output_path=output_path, sheet_name=None, theme=theme)


def upgrade_sheet_layout(workbook_path: str, output_path: str | None = None, *, sheet_name: str, theme: str = "corporate_formal") -> dict:
    return normalize_sheet_ui(workbook_path, output_path=output_path, sheet_name=sheet_name, theme=theme)
