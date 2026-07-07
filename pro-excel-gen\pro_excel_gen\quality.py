from __future__ import annotations

import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook


FORMULA_ERROR_RE = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A", re.I)
PLACEHOLDER_RE = re.compile(r"TODO|TBD|示例|占位|placeholder", re.I)


def count_media_parts(xlsx_path: str | Path) -> int:
    with zipfile.ZipFile(xlsx_path) as zf:
        return sum(1 for name in zf.namelist() if name.startswith("xl/media/"))


def count_chart_parts(xlsx_path: str | Path) -> int:
    with zipfile.ZipFile(xlsx_path) as zf:
        return sum(1 for name in zf.namelist() if name.startswith("xl/charts/") and name.endswith(".xml"))


def count_drawing_parts(xlsx_path: str | Path) -> int:
    with zipfile.ZipFile(xlsx_path) as zf:
        return sum(1 for name in zf.namelist() if name.startswith("xl/drawings/") and name.endswith(".xml"))


def inspect_workbook(xlsx_path: str | Path) -> dict:
    path = Path(xlsx_path)
    workbook = load_workbook(path, data_only=False)
    metrics = {
        "sheet_count": len(workbook.worksheets),
        "formula_count": 0,
        "placeholder_hits": 0,
        "formula_error_hits": 0,
        "table_count": 0,
        "media_count": count_media_parts(path),
        "chart_count": count_chart_parts(path),
        "drawing_count": count_drawing_parts(path),
        "sheet_names": workbook.sheetnames,
    }

    for ws in workbook.worksheets:
        metrics["table_count"] += len(getattr(ws, "tables", {}))
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if PLACEHOLDER_RE.search(value):
                        metrics["placeholder_hits"] += 1
                    if FORMULA_ERROR_RE.search(value):
                        metrics["formula_error_hits"] += 1
                if isinstance(value, str) and value.startswith("="):
                    metrics["formula_count"] += 1
                elif cell.data_type == "f":
                    metrics["formula_count"] += 1
    return metrics


def scan_formula_issues(xlsx_path: str | Path) -> dict:
    path = Path(xlsx_path)
    workbook = load_workbook(path, data_only=False)
    issues = []
    formula_count = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if FORMULA_ERROR_RE.search(value):
                        issues.append({"sheet": ws.title, "cell": cell.coordinate, "formula": value, "issue": "error_token"})
                    if "#REF!" in value.upper():
                        issues.append({"sheet": ws.title, "cell": cell.coordinate, "formula": value, "issue": "bad_reference"})
                elif isinstance(value, str) and FORMULA_ERROR_RE.search(value):
                    issues.append({"sheet": ws.title, "cell": cell.coordinate, "value": value, "issue": "visible_error"})
    return {"formula_count": formula_count, "issues": issues, "error_count": len(issues)}
