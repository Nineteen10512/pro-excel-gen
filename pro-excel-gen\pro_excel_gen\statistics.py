"""PRO-EXCEL v1.3.0 — 统计分析模块 (Batch 2: statistics)

合并自 spreadsheet-analyst 的深度统计分析能力:
- 增强描述统计 (std, quartiles, skew, kurtosis)
- 相关性矩阵
- 趋势分析 (移动平均, 同比环比)
- 分组聚合
- 透视分析
- 统计结果写入工作表
"""

from __future__ import annotations

from collections import defaultdict
from copy import deepcopy
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .delivery_kernel import import_tabular_data
from .data_cleaner import _detect_numeric_columns


# ── 增强描述统计 ────────────────────────────────────────────

def descriptive_stats(
    rows: list[dict],
    *,
    columns: list[str] | None = None,
) -> dict:
    """增强描述统计，含 std, quartiles, skew, kurtosis。

    Args:
        rows: 行数据列表
        columns: 限定统计列，None 为所有数值列

    Returns:
        {column_name: {count, sum, mean, std, min, q25, median, q75, max, skew, kurtosis}}
    """
    if not rows:
        return {}

    cols = columns or _detect_numeric_columns(rows)
    result = {}

    for col in cols:
        values = [row.get(col) for row in rows
                  if isinstance(row.get(col), (int, float)) and row.get(col) is not None]
        if len(values) < 2:
            continue

        arr = np.array(values, dtype=float)
        result[col] = {
            "count": len(values),
            "sum": float(np.sum(arr)),
            "mean": float(np.mean(arr)),
            "std": float(np.std(arr, ddof=1)),
            "min": float(np.min(arr)),
            "q25": float(np.percentile(arr, 25)),
            "median": float(np.median(arr)),
            "q75": float(np.percentile(arr, 75)),
            "max": float(np.max(arr)),
        }
        if len(values) >= 4:
            from scipy import stats
            result[col]["skew"] = float(stats.skew(arr))
            result[col]["kurtosis"] = float(stats.kurtosis(arr))

    return result


# ── 相关性矩阵 ──────────────────────────────────────────────

def correlation_analysis(
    rows: list[dict],
    *,
    columns: list[str] | None = None,
    method: str = "pearson",
) -> dict:
    """相关性矩阵分析。

    Args:
        rows: 行数据列表
        columns: 限定列，None 为所有数值列
        method: pearson | spearman

    Returns:
        {matrix: [[...]], columns: [...], pairs: [{col1, col2, correlation}]}
    """
    cols = columns or _detect_numeric_columns(rows)
    if len(cols) < 2:
        return {"matrix": [], "columns": cols, "pairs": []}

    # 提取数据矩阵
    data = []
    valid_cols = []
    for col in cols:
        values = [row.get(col) for row in rows
                  if isinstance(row.get(col), (int, float)) and row.get(col) is not None]
        if len(values) >= 3:
            data.append(values)
            valid_cols.append(col)

    if len(valid_cols) < 2:
        return {"matrix": [], "columns": valid_cols, "pairs": []}

    # 对齐长度
    min_len = min(len(d) for d in data)
    data = [d[:min_len] for d in data]
    arr = np.array(data)

    if method == "spearman":
        from scipy.stats import spearmanr
        corr_matrix = np.zeros((len(valid_cols), len(valid_cols)))
        for i in range(len(valid_cols)):
            for j in range(len(valid_cols)):
                if i == j:
                    corr_matrix[i][j] = 1.0
                elif j > i:
                    corr, _ = spearmanr(arr[i], arr[j])
                    corr_matrix[i][j] = corr
                    corr_matrix[j][i] = corr
    else:
        corr_matrix = np.corrcoef(arr)

    # 构建 pairs
    pairs = []
    for i in range(len(valid_cols)):
        for j in range(i + 1, len(valid_cols)):
            pairs.append({
                "col1": valid_cols[i],
                "col2": valid_cols[j],
                "correlation": round(float(corr_matrix[i][j]), 4),
            })

    return {
        "matrix": [[round(float(v), 4) for v in row] for row in corr_matrix],
        "columns": valid_cols,
        "pairs": sorted(pairs, key=lambda x: abs(x["correlation"]), reverse=True),
    }


# ── 趋势分析 ────────────────────────────────────────────────

def trend_analysis(
    rows: list[dict],
    *,
    date_column: str,
    value_column: str,
    window: int = 3,
) -> dict:
    """趋势分析：移动平均 + 环比变化。

    Args:
        rows: 行数据列表
        date_column: 日期列名
        value_column: 数值列名
        window: 移动平均窗口大小

    Returns:
        {moving_average: [...], period_change: [...], trend_direction: str}
    """
    if not rows:
        return {"moving_average": [], "period_change": [], "trend_direction": "flat"}

    values = []
    for row in rows:
        val = row.get(value_column)
        if isinstance(val, (int, float)):
            values.append(float(val))

    if len(values) < 2:
        return {"moving_average": [], "period_change": [], "trend_direction": "flat"}

    # 移动平均
    arr = np.array(values)
    ma = []
    for i in range(len(arr)):
        start = max(0, i - window + 1)
        ma.append(round(float(np.mean(arr[start:i + 1])), 2))

    # 环比变化
    changes = []
    for i in range(1, len(arr)):
        if arr[i - 1] != 0:
            pct = round(float((arr[i] - arr[i - 1]) / abs(arr[i - 1]) * 100), 2)
        else:
            pct = None
        changes.append(pct)

    # 趋势方向
    if len(arr) >= 3:
        slope = np.polyfit(range(len(arr)), arr, 1)[0]
        if slope > 0.01 * np.mean(arr):
            direction = "up"
        elif slope < -0.01 * np.mean(arr):
            direction = "down"
        else:
            direction = "flat"
    else:
        direction = "flat"

    return {
        "moving_average": ma,
        "period_change": changes,
        "trend_direction": direction,
        "total_change_pct": round(float((arr[-1] - arr[0]) / abs(arr[0]) * 100), 2) if arr[0] != 0 else None,
    }


# ── 分组聚合 ────────────────────────────────────────────────

def group_aggregate(
    rows: list[dict],
    *,
    group_by: str | list[str],
    aggregations: dict[str, str | list[str]],
) -> list[dict]:
    """分组聚合。

    Args:
        rows: 行数据列表
        group_by: 分组列
        aggregations: {列名: 聚合函数}，如 {"销售额": "sum", "利润": ["mean", "max"]}
            支持: sum, mean, median, min, max, count, std

    Returns:
        聚合后的行列表
    """
    if not rows:
        return []

    groups = group_by if isinstance(group_by, list) else [group_by]
    buckets = defaultdict(list)

    for row in rows:
        key = tuple(str(row.get(g)) for g in groups)
        buckets[key].append(row)

    result = []
    for key, bucket_rows in buckets.items():
        agg_row = {}
        for i, g in enumerate(groups):
            agg_row[g] = key[i]

        for col, funcs in aggregations.items():
            func_list = funcs if isinstance(funcs, list) else [funcs]
            values = [row.get(col) for row in bucket_rows
                      if isinstance(row.get(col), (int, float)) and row.get(col) is not None]

            for func in func_list:
                col_name = f"{col}_{func}"
                if not values:
                    agg_row[col_name] = None
                    continue

                arr = np.array(values, dtype=float)
                if func == "sum":
                    agg_row[col_name] = round(float(np.sum(arr)), 2)
                elif func == "mean":
                    agg_row[col_name] = round(float(np.mean(arr)), 2)
                elif func == "median":
                    agg_row[col_name] = round(float(np.median(arr)), 2)
                elif func == "min":
                    agg_row[col_name] = round(float(np.min(arr)), 2)
                elif func == "max":
                    agg_row[col_name] = round(float(np.max(arr)), 2)
                elif func == "count":
                    agg_row[col_name] = len(values)
                elif func == "std":
                    agg_row[col_name] = round(float(np.std(arr, ddof=1)), 2) if len(values) >= 2 else None

        result.append(agg_row)

    return result


# ── 透视分析 ────────────────────────────────────────────────

def pivot_analysis(
    rows: list[dict],
    *,
    index: str,
    columns: str,
    values: str,
    aggfunc: str = "sum",
) -> dict:
    """透视分析。

    Args:
        rows: 行数据列表
        index: 行索引列
        columns: 列索引列
        values: 值列
        aggfunc: 聚合函数 (sum, mean, count)

    Returns:
        {pivot_table: [[...]], row_labels: [...], col_labels: [...]}
    """
    if not rows:
        return {"pivot_table": [], "row_labels": [], "col_labels": []}

    # 收集唯一值
    row_labels = sorted(set(str(row.get(index)) for row in rows))
    col_labels = sorted(set(str(row.get(columns)) for row in rows))

    # 构建查找
    lookup = {}
    for row in rows:
        rk = str(row.get(index))
        ck = str(row.get(columns))
        val = row.get(values)
        if isinstance(val, (int, float)):
            key = (rk, ck)
            if key not in lookup:
                lookup[key] = []
            lookup[key].append(val)

    # 构建透视表
    table = []
    for rl in row_labels:
        row_data = []
        for cl in col_labels:
            vals = lookup.get((rl, cl), [])
            if not vals:
                row_data.append(None)
            elif aggfunc == "sum":
                row_data.append(round(sum(vals), 2))
            elif aggfunc == "mean":
                row_data.append(round(np.mean(vals), 2))
            elif aggfunc == "count":
                row_data.append(len(vals))
        table.append(row_data)

    return {
        "pivot_table": table,
        "row_labels": row_labels,
        "col_labels": col_labels,
    }


# ── 统计结果写入工作表 ──────────────────────────────────────

def add_statistics_sheet(
    workbook_path: str,
    output_path: str | None = None,
    *,
    sheet_name: str | None = None,
    theme: str = "corporate_formal",
) -> dict:
    """将增强统计结果写入新工作表。

    Args:
        workbook_path: 源工作簿路径
        output_path: 输出路径，None 则覆盖
        sheet_name: 数据源工作表名
        theme: 主题

    Returns:
        {output_path, sheets_added: [...]}
    """
    target = output_path or workbook_path
    rows = import_tabular_data(workbook_path, sheet_name=sheet_name)

    if not rows:
        return {"output_path": target, "sheets_added": []}

    stats = descriptive_stats(rows)
    corr = correlation_analysis(rows)

    # 风格
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    workbook = load_workbook(workbook_path)
    sheets_added = []

    # 描述统计 sheet
    if "Statistics" in workbook.sheetnames:
        del workbook["Statistics"]
    ws = workbook.create_sheet("Statistics")

    stat_headers = ["指标", "count", "mean", "std", "min", "q25", "median", "q75", "max", "skew", "kurtosis"]
    for c, h in enumerate(stat_headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for r, (col, s) in enumerate(stats.items(), 2):
        ws.cell(r, 1, col).border = thin_border
        for c, key in enumerate(stat_headers[1:], 2):
            val = s.get(key)
            cell = ws.cell(r, c, round(val, 4) if isinstance(val, float) else val)
            cell.border = thin_border

    sheets_added.append("Statistics")

    # 相关性矩阵 sheet
    if corr["pairs"]:
        if "Correlation" in workbook.sheetnames:
            del workbook["Correlation"]
        ws2 = workbook.create_sheet("Correlation")

        for c, col in enumerate(corr["columns"], 1):
            cell = ws2.cell(1, c, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for r, pair in enumerate(corr["pairs"], 2):
            ws2.cell(r, 1, pair["col1"]).border = thin_border
            ws2.cell(r, 2, pair["col2"]).border = thin_border
            cell = ws2.cell(r, 3, pair["correlation"])
            cell.border = thin_border
            cell.number_format = "0.0000"

        sheets_added.append("Correlation")

    workbook.save(target)

    from .normalizer import normalize_workbook_ui
    normalize_workbook_ui(target, theme=theme)

    return {"output_path": target, "sheets_added": sheets_added, "stats_columns": len(stats), "corr_pairs": len(corr["pairs"])}